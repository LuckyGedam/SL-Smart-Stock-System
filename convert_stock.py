"""
S L Crockeries Stock Report → SSPS Format
Stock = closing (actual pieces in godown)
"""
import re, csv, sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    import subprocess; subprocess.run([sys.executable,"-m","pip","install","openpyxl","-q"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

FILTERS = {
    "only_positive_stock" : True,
    "min_stock"           : 1,
    "max_stock"           : None,
    "max_age_days"        : None,
    "include_categories"  : [],
    "include_colors"      : [],
    "name_keywords"       : [],
    "min_price"           : None,
    "max_price"           : None,
}

def parse_date(val) -> str:
    if isinstance(val, datetime): return val.strftime("%Y-%m-%d")
    s = str(val or "").strip()[:10]
    return s if len(s) == 10 else datetime.today().strftime("%Y-%m-%d")

def to_int(val, default=0) -> int:
    try: return int(float(str(val or 0)))
    except: return default

def to_float(val, default=0.0) -> float:
    try: return float(str(val or 0))
    except: return default

def clean(val) -> str:
    return str(val or "").strip()

def guess_category(name: str, barcode: str = "") -> str:
    n = name.upper()
    if any(w in n for w in ["MUG","CUP","COFFEE MUG"]): return "Mug / Cup"
    if any(w in n for w in ["GLASS","GOBLET","TUMBLER","WHISKY","WINE","BEER","GIMLET","SHOT"]): return "Glassware"
    if any(w in n for w in ["BOWL"]): return "Bowl"
    if any(w in n for w in ["DINNER SET","DINNER","THALI","PLATE"]): return "Dinner Set"
    if any(w in n for w in ["JUG","PITCHER","DECANTER","DISPENSER","DESPENSER","CASSEROLE"]): return "Jug"
    if any(w in n for w in ["SPOON","FORK","KNIFE","CUTLERY"]): return "Cutlery"
    if any(w in n for w in ["TEA SET","KETTLE","TEA POT"]): return "Tea Set"
    if "STEEL" in n: return "Steel"
    if "PLASTIC" in n: return "Plastic"
    p = str(barcode)[:3]
    return {"361":"Glassware","261":"Crockery","461":"Cutlery","561":"Plastic","661":"Steel","761":"Melamine"}.get(p,"General")

def apply_filters(row: dict, is_catalog=False) -> bool:
    stock = row.get("stock", 0)
    price = row.get("price", 0.0)
    color = (row.get("color") or "").upper()
    name  = (row.get("name")  or "").upper()
    cat   = row.get("category", "")
    age   = row.get("age_days", 0)
    if not is_catalog:
        if FILTERS["only_positive_stock"] and stock < FILTERS["min_stock"]: return False
        if FILTERS["max_stock"] is not None and stock > FILTERS["max_stock"]: return False
    if FILTERS["max_age_days"] is not None and age > FILTERS["max_age_days"]: return False
    if FILTERS["include_categories"] and cat not in FILTERS["include_categories"]: return False
    if FILTERS["include_colors"] and color not in [c.upper() for c in FILTERS["include_colors"]]: return False
    if FILTERS["name_keywords"] and not any(kw.upper() in name for kw in FILTERS["name_keywords"]): return False
    if FILTERS["min_price"] is not None and price < FILTERS["min_price"]: return False
    if FILTERS["max_price"] is not None and price > FILTERS["max_price"]: return False
    return True

def parse_old_excel(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    header_row_num = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and str(row[0]).strip().lower() == "barcodenumber":
            header_row_num = i; break
    if not header_row_num:
        raise RuntimeError(f"Cannot find header row in {filepath}")
    headers = [str(c or "").strip() for c in
               list(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))[0]]
    col = {h: i for i, h in enumerate(headers)}

    products = []
    for idx, row in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=True), 1):
        if not row or row[0] is None: continue
        barcode = clean(row[col.get("BarcodeNumber", 0)])
        if not barcode or barcode.lower() in ("total","barcodenumber"): continue

        raw_name  = clean(row[col.get("ItemName", 20)])
        brand     = clean(row[col.get("Brand", 25)])
        itype     = clean(row[col.get("Type", 27)])
        color     = clean(row[col.get("Color", 24)])
        # ✅ FIXED: Use 'closing' for actual piece count, NOT MrpStockValue
        closing   = row[col.get("closing", 4)]
        age_days  = to_int(row[col.get("Age", 7)])
        pur_date  = row[col.get("PurDate", 6)]
        mrp_rate  = row[col.get("MrpRate", 8)]

        sku_match = re.search(r'(HL-?\d+|\d{5,})', raw_name, re.I)
        sku = sku_match.group(1).upper() if sku_match else f"HL-{barcode[-6:]}"
        alpha = re.sub(r'(HL-?\d+|\d{4,})', '', raw_name).strip(" .-")
        if len(alpha) > 2:
            product_name = alpha
        elif brand:
            product_name = f"{brand} {itype}".strip()
        else:
            product_name = f"Item {barcode}"

        # closing can be negative (means sold more than received), use abs or 0
        stock = max(0, to_int(closing))

        products.append({
            "sku"       : sku,
            "name"      : product_name,
            "category"  : guess_category(product_name + " " + raw_name, barcode),
            "stock"     : stock,
            "date_added": parse_date(pur_date),
            "barcode"   : barcode,
            "color"     : color,
            "item_type" : itype,
            "age_days"  : age_days,
            "price"     : to_float(mrp_rate),
        })
    return products

def parse_new_csv(filepath: str) -> tuple[list[dict], bool]:
    with open(filepath, encoding="utf-8", errors="ignore") as f:
        rows = list(csv.DictReader(f))
    filled = sum(1 for r in rows if r.get("Available quantity","").strip())
    has_stock = filled > len(rows) * 0.05
    products = []
    for idx, r in enumerate(rows, 1):
        name = clean(r.get("Product Name",""))
        if not name: continue
        raw_sku  = clean(r.get("Sku",""))
        pid      = clean(r.get("Product id",""))
        color    = clean(r.get("Colours") or r.get("colour") or r.get("colors") or "")
        qty_raw  = clean(r.get("Available quantity",""))
        stock    = to_int(qty_raw) if qty_raw else 0
        price    = to_float(r.get("Product Price") or r.get("Discounted price") or 0)
        pcs      = clean(r.get("Pcs") or r.get("Set Name") or "")
        sku = re.sub(r'\s+','_', raw_sku.upper()[:20]) if raw_sku else f"SLP{idx:04d}"
        barcode = raw_sku if raw_sku else pid
        pcs_match = re.search(r'(\d+\s*PCS|\d+\s*PC\.?)', name, re.I)
        item_type = pcs if pcs else (pcs_match.group(1) if pcs_match else "")
        products.append({
            "sku": sku, "name": name,
            "category": guess_category(name),
            "stock": stock,
            "date_added": datetime.today().strftime("%Y-%m-%d"),
            "barcode": barcode, "color": color,
            "item_type": item_type, "age_days": 0, "price": price,
        })
    return products, not has_stock

CAT_COLORS = {
    "Glassware":"DBEAFE","Mug / Cup":"FED7AA","Bowl":"BBF7D0",
    "Dinner Set":"E9D5FF","Jug":"FEF08A","Crockery":"BAE6FD",
    "Cutlery":"FECACA","Tea Set":"A7F3D0","Steel":"E2E8F0",
    "Plastic":"F5F5F4","Melamine":"FCE7F3",
}

def write_output(products: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SSPS Products"
    # ✅ Columns backend expects
    COLS = ["SKU","Product Name","Category","Stock","Added Date","Barcode","Color","Item Type","Age Days","Price (MRP)"]
    ws.append(COLS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="1E293B")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in products:
        ws.append([
            row["sku"], row["name"], row["category"], row["stock"],
            row["date_added"], row["barcode"], row["color"],
            row["item_type"], row["age_days"], row["price"],
        ])
        fill = PatternFill("solid", fgColor=CAT_COLORS.get(row.get("category",""), "FFFFFF"))
        for cell in ws[ws.max_row]:
            cell.fill = fill
    widths = [16,32,20,8,13,16,10,12,10,12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)

import tkinter as tk
from tkinter import filedialog, messagebox

if __name__ == "__main__":
    root = tk.Tk(); root.withdraw(); root.attributes("-topmost", True)
    input_files = filedialog.askopenfilenames(
        title="Select Excel / CSV files",
        filetypes=[("Excel & CSV","*.xlsx *.xls *.csv"),("All files","*.*")]
    )
    if not input_files:
        messagebox.showinfo("Cancelled","No files selected."); sys.exit(0)
    output_dir = filedialog.askdirectory(title="Select OUTPUT folder")
    if not output_dir:
        messagebox.showinfo("Cancelled","No output folder selected."); sys.exit(0)
    output_dir = Path(output_dir)
    summary = []
    for inp in input_files:
        inp_path = Path(inp)
        out_path = output_dir / f"SSPS_{inp_path.stem}.xlsx"
        ext = inp_path.suffix.lower()
        print(f"\n📂 Reading: {inp_path.name}")
        try:
            if ext in (".xlsx",".xls"):
                products = parse_old_excel(inp); is_catalog = False
            elif ext == ".csv":
                products, is_catalog = parse_new_csv(inp)
                if is_catalog: print("   ℹ️  Catalog mode")
            else:
                continue
            total = len(products)
            filtered = [p for p in products if apply_filters(p, is_catalog=is_catalog)]
            write_output(filtered, str(out_path))
            msg = f"{inp_path.name} → {len(filtered)} products saved"
            print(f"✅ {msg}")
            summary.append(msg)
        except Exception as e:
            err = f"Error in {inp_path.name}: {e}"
            print(f"❌ {err}"); summary.append(f"❌ {err}")
    messagebox.showinfo("Done ✅", "\n".join(summary))