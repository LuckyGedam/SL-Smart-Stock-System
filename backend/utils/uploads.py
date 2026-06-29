import io
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl import load_workbook

from ..schemas.product import ProductCreate


EXPECTED_COLUMNS = ["SKU", "Product Name", "Category", "Stock", "Added Date", "Barcode", "url"]


# All column aliases we accept
COLUMN_MAP = {
    "sku"        : "SKU",
    "productname": "Product Name",
    "category"   : "Category",
    "stock"      : "Stock",
    "addeddate"  : "Added Date",
    "barcode"    : "Barcode",
    "url"        : "url",

    # extras from our converter output (ignored, not required)
    "color"      : None,
    "itemtype"   : None,
    "agedays"    : None,
    "pricemrp"   : None,
    "price(mrp)" : None,
}

def _normalize(s: str) -> str:
    return str(s).lower().replace(" ", "").replace("_", "").replace("-", "").replace("(", "").replace(")", "")

def read_excel_rows(file_bytes: bytes) -> list[dict[str, Any]]:
    import traceback

    print("[DEBUG] Excel parser started")
    workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    print(f"[DEBUG] Excel rows read (including header/empty): {len(rows)}")

    if not rows:
        return []

    # Find actual header row (first row with 'SKU' or 'sku' in any cell)
    header_row_idx = 0
    for i, row in enumerate(rows[:10]):
        normalized = [_normalize(str(c or "")) for c in row]
        if "sku" in normalized:
            header_row_idx = i
            break

    raw_headers = rows[header_row_idx]
    headers = [str(cell).strip() if cell is not None else "" for cell in raw_headers]
    normalized_headers = [_normalize(h) for h in headers]

    # Check required columns present (loose match, any order)
    required = ["sku", "productname", "category", "stock", "addeddate", "barcode"]
    # "url" column is optional (image_url); if missing we'll keep it empty.

    missing = [r for r in required if r not in normalized_headers]
    if missing:
        raise ValueError(f"Missing required columns: {missing}. Got: {headers}")

    # Build index map
    col_idx = {_normalize(h): i for i, h in enumerate(headers) if h}

    products: list[dict[str, Any]] = []

    parsed_first = False
    for row in rows[header_row_idx + 1:]:
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue


        def get(key: str):
            idx = col_idx.get(key)
            return row[idx] if idx is not None and idx < len(row) else None

        sku  = str(get("sku") or "").strip()
        name = str(get("productname") or "").strip()

        if not sku or not name:
            continue

        # Handle Excel date objects in Added Date
        date_val = get("addeddate")
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val or "").strip()[:10]

        try:
            stock = int(float(str(get("stock") or 0)))
        except (ValueError, TypeError):
            stock = 0

        try:
            price = float(str(get("pricemrp") or get("price(mrp)") or 0))
        except (ValueError, TypeError):
            price = 0.0

        url_val = get("url")
        url_str = str(url_val or "").strip() if url_val is not None else ""

        product_dict = {
            "sku"          : sku,
            "name"         : name,

            "category"     : str(get("category") or "General").strip(),
            "currentStock" : stock,
            "dateAdded"    : date_str,
            "barcode"      : str(get("barcode") or "").strip(),
            "price"        : price,
            "image_url"    : url_str,
            # legacy image field (kept in sync)
            "image"         : url_str,

            "location"      : "",

        }

        if not parsed_first:
            print(f"[DEBUG] Parsed excel row (first)={product_dict}")
            parsed_first = True

        products.append(product_dict)

    print(f"[DEBUG] Parsed {len(products)} product dict rows")
    return products



def build_product_payloads(rows: list[dict[str, Any]]) -> list[ProductCreate]:
    return [ProductCreate(**row) for row in rows]
